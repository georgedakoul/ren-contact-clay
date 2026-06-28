"""discover_contacts.py — Clay contact discovery driven by actionable_contacts.json.

Usage:
  python discover_contacts.py                          → print status report
  python discover_contacts.py save                     → save the BATCH dict below
  python discover_contacts.py enrich_emails Brand1 … → print JSON for find-and-enrich-list-of-contacts
  python discover_contacts.py export                   → export all contacts to Excel
  python discover_contacts.py mark_empty   Brand1 [Brand2 ...]
  python discover_contacts.py unmark_empty Brand1 [Brand2 ...]

3-phase Clay workflow:
  Phase 1 — Employee discovery:
    find-and-enrich-contacts-at-company(companyIdentifier=<domain>, contactFilters={locations:["Greece"]})
    → raw employees (name + title + LinkedIn URL, no email needed yet)
  Phase 2 — Save + title filter:
    Edit BATCH dict, run `python discover_contacts.py save`
    → writes 00-BrandName.json, BANNED_TITLES applied automatically
  Phase 3 — Email enrichment:
    run `python discover_contacts.py enrich_emails BrandName`
    → prints JSON list; pass to find-and-enrich-list-of-contacts with dataPoints:{contactDataPoints:[{type:"Email"}]}
    → save result via BATCH as usual

File prefix convention in output/employees/:
  BrandName.json    → ✓ Covered   (has at least one email)
  00-BrandName.json → ∅ LinkedIn  (contacts found, no emails yet)
  ZZ-BrandName.json → ✗ Empty     (Clay returned 0 contacts — skip)
  (no file)         → ✗ Missing   (never searched)
"""
import json, sys, unicodedata, re, subprocess
from datetime import datetime, timezone
from pathlib import Path

ROOT          = Path(__file__).resolve().parents[3]
STORE_DIR     = ROOT / "AI Sales Agent System" / "output" / "employees"
CONTACTS_FILE = ROOT / "AI Sales Agent System" / "actionable_contacts.json"
EXPORT_PATH   = ROOT / "AI Sales Agent System" / "output" / "contacts_export.xlsx"
TODAY         = datetime.now(timezone.utc).strftime("%Y-%m-%d")
STORE_DIR.mkdir(parents=True, exist_ok=True)

# ---------------------------------------------------------------------------
# DOMAIN_MAP: preferred companyIdentifier (domain) per brand.
# Always use domain over LinkedIn slug — Clay resolves the correct entity from domain.
# Add any brand here once its domain is known; brands not listed fall back to linkedin_slug.
# ---------------------------------------------------------------------------
DOMAIN_MAP = {
    # Slug resolves to wrong entity or returns 0 — domain is the fix
    "COSMOTE":           "cosmote.gr",
    "Village Cinemas":   "villagecinemas.gr",    # slug → Australian entity
    "Cosmos Sport":      "cosmossport.gr",
    "Coca-Cola":         "coca-colahellenic.com",
    "Psichogios Books":  "psichogios.gr",
    "Pame Stoixima":     "opap.gr",              # brand lives under OPAP
    "ION":               "ion.gr",
    "instacar":          "instacar.gr",          # slug → US company
    "Apivita":           "apivita.com",
    "more.com":          "more.com",
    "Wind":              "wind.gr",
    "Alterlife":         "alterlife.gr",
    "SKY express":       "skyexpress.gr",
    "Carroten":          "carroten.gr",          # owned by Sarantis; no standalone LinkedIn
    "Fresh Line":        "freshline.gr",
    "Alumil":            "alumil.com",
    "BSB Fashion":       "bsbfashion.com",
    "La Vie en Rose":    "lavieenrose.com",      # slug → Swiss NGO
    "Mind Your Style":   "mindyourstyle.gr",
    "Protergia":         "protergia.gr",
    # Known good domains from previous batches
    "Douleutaras":       "douleutaras.gr",
    "Box Now":           "boxnow.gr",
    "Germanos":          "germanos.gr",
    "MEVGAL":            "mevgal.gr",
    "Vitex":             "vitex.gr",
    "Three Cents":       "threecents.com",
    "Snappi":            "snappibank.com",
    "LG":                "lg.com",
    "JYSK":              "jysk.com",
    "Kinder":            "ferrerocareers.com",
    "NIVEA":             "beiersdorf.com",
    "Dove":              "unilever.com",
    "Converse":          "converse.com",
    "BMW":               "bmwgroup.com",
    "Ferryhopper":       "ferryhopper.com",
    "Vans":              "vans.com",
    "ANT1":              "ant1.gr",
    "AEK FC":            "aekfc.gr",
    "Jumbo":             "e-jumbo.gr",
    "FAGE":              "home.fage",
    "New Balance":       "newbalance.com",
    "Zara":              "inditexpeople.com",
    "Puma":              "puma.com",
    "CarVertical":       "carvertical.com",
}

# All searches now use locations=["Greece"] — this is a Greek outreach list.
# GLOBAL_BRANDS kept for reference only (no longer drives filter logic).
GLOBAL_BRANDS = {
    "Samsung", "Apple", "Herbalife", "IKEA", "LEGO", "Huawei",
    "Motorola", "Starbucks", "Red Bull", "Wolt", "FREENOW",
    "Hertz", "Sony Music Entertainment", "Chanel", "Visa",
    "Nespresso", "Monster Energy", "La Vie en Rose",
}

# Job titles to ignore — contacts whose title contains any of these strings
# (case-insensitive, substring match) are skipped on save and purged on export.
BANNED_TITLES = {
    # Confirmed by user (specific titles)
    "chief of staff to the ceo",
    "design lead",
    "performance marketing manager",
    "assistant general manager",
    "chief executive officer at cosmote payments",   # specific person only, not all CEOs
    # Business development / commercial (not marketing decision-makers)
    "business development",         # BDM, head of BD, chief strategy & BD officer
    "commercial director",
    "commercial manager",
    "international business director",
    # Trade / shopper / channel (not influencer marketing decision-makers)
    "trade marketing",              # specialist, coordinator, lead, section manager, manager
    "shopper",                      # shopper marketing manager, head of shopper marketing
    # Exports (distribution, not brand marketing)
    "exports director",
    "export marketing manager",
    # Operations
    "director of partners operations",
    # HR / people / admin roles
    "head of people",               # head of people rewards, head of people & culture, etc.
    "organisational development",
    "executive assistant",
    # Analytics (not decision-makers)
    "business data analyst",
    # Keyword catch-all
    "sales",                        # any title containing "sales"
    # Noise / non-decision-maker
    "marketing department",         # not a real title
    "marketing team member",        # too junior / generic
    "jysk influencer",              # brand ambassador role, not marketing decision-maker
    "assistant brand marketing skip",  # support role
    # FLA-specific / internal codes
    "fla marketing director",
    # Performance/paid specialists that are executional, not decision-makers
    "performance marketing specialist",
}

# Exact-match bans (full title only — substring would catch valid roles).
BANNED_TITLES_EXACT = {
    "social media",                 # bare "Social Media" with no qualifier is noise
}

# Titles always kept regardless of BANNED_TITLES substring matches.
# Paid media roles are valid decision-makers for influencer spend conversations.
WANTED_TITLES = {
    "paid media manager",
    "paid media specialist",
}

_brand_names_cache = None

def _brand_names():
    global _brand_names_cache
    if _brand_names_cache is None:
        try:
            data = json.loads(CONTACTS_FILE.read_text(encoding="utf-8"))
            _brand_names_cache = {normalize(b["brand"]) for b in data["brands"]}
        except Exception:
            _brand_names_cache = set()
    return _brand_names_cache


def _is_banned(title):
    t = normalize(title)
    if t in BANNED_TITLES_EXACT:
        return True
    if t in WANTED_TITLES:
        return False
    # Brand-prefixed titles (e.g. "BMW Marketing Manager") always kept.
    for brand in _brand_names():
        if t.startswith(brand + " "):
            return False
    return any(banned in t for banned in BANNED_TITLES)


def _strip_accents(s):
    return "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))

def normalize(s):
    return re.sub(r"\s+", " ", _strip_accents(s).lower().strip())


def get_identifier(brand_name, linkedin_slug, website=None):
    if brand_name in DOMAIN_MAP:
        return DOMAIN_MAP[brand_name]
    if website:
        return website
    if linkedin_slug:
        return f"https://www.linkedin.com/company/{linkedin_slug}"
    return None  # no identifier — needs domain discovery via web search


def _employee_path(brand_name):
    """Return existing employee file; checks all prefix variants."""
    for prefix in ("", "00-", "ZZ-"):
        p = STORE_DIR / f"{prefix}{brand_name}.json"
        if p.exists():
            return p
    return STORE_DIR / f"{brand_name}.json"  # default for new files


def _sync_actionable(brand_name):
    """Update actionable_contacts.json stats for one brand from the employee file on disk."""
    try:
        data = json.loads(CONTACTS_FILE.read_text(encoding="utf-8"))
    except Exception:
        return
    brands = data if isinstance(data, list) else data.get("brands", [])
    entry = next((b for b in brands if b.get("brand") == brand_name), None)
    if entry is None:
        return
    info = _scan_store().get(brand_name)
    if info is None:
        entry["has_employee_file"] = False
        entry["employee_count"]    = 0
        entry["emails_known"]      = 0
        entry["employee_file"]     = None
        entry["coverage_status"]   = "No coverage"
    elif info["_prefix"] == "ZZ-":
        entry["has_employee_file"] = True
        entry["employee_count"]    = 0
        entry["emails_known"]      = 0
        entry["employee_file"]     = f"ZZ-{brand_name}.json"
        entry["coverage_status"]   = "Empty"
    elif info["emails"] > 0:
        entry["has_employee_file"] = True
        entry["employee_count"]    = info["total"]
        entry["emails_known"]      = info["emails"]
        entry["employee_file"]     = f"{brand_name}.json"
        entry["coverage_status"]   = "Covered"
    else:
        entry["has_employee_file"] = True
        entry["employee_count"]    = info["total"]
        entry["emails_known"]      = 0
        entry["employee_file"]     = f"00-{brand_name}.json"
        entry["coverage_status"]   = "LinkedIn-only"
    CONTACTS_FILE.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")


def save_contacts(brand_name, contacts, domain=None):
    if not contacts:
        zz = STORE_DIR / f"ZZ-{brand_name}.json"
        zz.write_text("[]", encoding="utf-8")
        for old_prefix in ("00-",):
            old = STORE_DIR / f"{old_prefix}{brand_name}.json"
            if old.exists():
                old.unlink()
        _sync_actionable(brand_name)
        print(f"  {brand_name}: Clay returned 0 contacts → ZZ-{brand_name}.json (skipped in future batches)")
        return
    # If we now have contacts, remove any ZZ- marker
    zz = STORE_DIR / f"ZZ-{brand_name}.json"
    if zz.exists():
        zz.unlink()
    path = _employee_path(brand_name)
    existing = json.loads(path.read_text(encoding="utf-8")) if path.exists() else []
    by_name = {normalize(e["name"]): e for e in existing}
    # Purge any existing contacts with banned titles before merging
    by_name = {k: v for k, v in by_name.items() if not _is_banned(v.get("job_title") or "")}
    added = email_added = 0
    for c in contacts:
        name = (c.get("name") or "").strip()
        if not name:
            continue
        title  = (c.get("latest_experience_title") or "").strip()
        if _is_banned(title):
            continue
        key = normalize(name)
        li_url = c.get("url") or None
        em     = (c.get("email") or "").strip() or None
        dom    = c.get("domain") or domain or None
        if key in by_name:
            if title  and not by_name[key].get("job_title"):    by_name[key]["job_title"]    = title
            if li_url and not by_name[key].get("linkedin_url"): by_name[key]["linkedin_url"] = li_url
            if em     and not by_name[key].get("email"):
                by_name[key]["email"] = em
                email_added += 1
            by_name[key]["last_seen"] = TODAY
        else:
            by_name[key] = {
                "name": name, "job_title": title or None, "email": em,
                "linkedin_url": li_url, "verified": False, "active": True,
                "source": "clay", "domain": dom, "first_seen": TODAY, "last_seen": TODAY,
            }
            added += 1
            if em: email_added += 1
    merged = list(by_name.values())
    has_email = any(c.get("email") for c in merged)
    new_path = STORE_DIR / (f"{brand_name}.json" if has_email else f"00-{brand_name}.json")
    new_path.write_text(json.dumps(merged, indent=2, ensure_ascii=False), encoding="utf-8")
    if path.exists() and path != new_path:
        path.unlink()
    _sync_actionable(brand_name)
    print(f"  {brand_name}: {added} new contacts, {email_added} emails added → {len(merged)} total")


def _scan_store():
    """Scan ALL employee files → {brand_name: {total, emails, last_seen, _prefix}}."""
    store = {}
    for f in sorted(STORE_DIR.glob("*.json")):
        if f.name == "desktop.ini":
            continue
        stem = f.stem
        if stem.startswith("ZZ-"):
            brand_name = stem[3:]
            store[brand_name] = {"total": 0, "emails": 0, "last_seen": None, "_prefix": "ZZ-"}
        elif stem.startswith("00-"):
            brand_name = stem[3:]
            contacts = json.loads(f.read_text(encoding="utf-8"))
            email_count = sum(1 for c in contacts if c.get("email"))
            last_seen = max((c.get("last_seen") or "1970-01-01") for c in contacts) if contacts else None
            store[brand_name] = {"total": len(contacts), "emails": email_count, "last_seen": last_seen, "_prefix": "00-"}
        else:
            brand_name = stem
            contacts = json.loads(f.read_text(encoding="utf-8"))
            email_count = sum(1 for c in contacts if c.get("email"))
            last_seen = max((c.get("last_seen") or "1970-01-01") for c in contacts) if contacts else None
            store[brand_name] = {"total": len(contacts), "emails": email_count, "last_seen": last_seen, "_prefix": ""}
    return store


def status_report():
    store = _scan_store()
    total_contacts = sum(v["total"] for v in store.values())
    total_emails   = sum(v["emails"] for v in store.values())
    brands_with_emails = sum(1 for v in store.values() if v["emails"] > 0)

    print(f"\n=== EMPLOYEE STORE OVERVIEW ===")
    print(f"  Files on disk : {len(store):>4}  ({brands_with_emails} with emails)")
    print(f"  Total contacts: {total_contacts:>4}  ({total_emails} with emails)")

    data      = json.loads(CONTACTS_FILE.read_text(encoding="utf-8"))
    ac_brands = {b["brand"]: b for b in data["brands"]}
    searchable = ac_brands

    covered      = []  # BrandName.json — has ≥1 email
    linkedin_only = [] # 00-BrandName.json — contacts but 0 emails
    empty        = []  # ZZ-BrandName.json — Clay returned 0 contacts
    missing      = []  # no file — has identifier (slug/override/website), never searched
    needs_domain = []  # no file — no identifier at all, needs web search to find domain

    for name in sorted(searchable):
        slug       = searchable[name].get("linkedin_slug") or ""
        website    = searchable[name].get("website") or None
        identifier = get_identifier(name, slug, website)
        info       = store.get(name)
        if identifier is None:
            needs_domain.append((name,))
        elif info is None:
            missing.append((name, identifier))
        elif info["_prefix"] == "ZZ-":
            empty.append((name, identifier))
        elif info["emails"] > 0:
            covered.append((name, identifier, info))
        else:
            linkedin_only.append((name, identifier, info))

    print(f"\n=== CLAY-SEARCHABLE BRANDS ({len(searchable)} total) ===")
    print(f"  ✓ Covered       (BrandName.json   — has emails)        : {len(covered)}")
    print(f"  ∅ LinkedIn-only (00-BrandName.json — contacts, 0 email): {len(linkedin_only)}")
    print(f"  ✗ Empty         (ZZ-BrandName.json — Clay returned 0)  : {len(empty)}")
    print(f"  ✗ Missing       (has identifier    — never searched)   : {len(missing)}")
    print(f"  ? Needs-domain  (no identifier     — web search needed): {len(needs_domain)}")

    if covered:
        print(f"\n--- ✓ COVERED (have emails) ---")
        print(f"  {'Brand':<32} {'Identifier':<45} {'#C':>4} {'#E':>4} {'Last seen':<12}")
        print("  " + "─" * 100)
        for name, identifier, info in covered:
            print(f"  {name:<32} {identifier:<45} {info['total']:>4} {info['emails']:>4} {info['last_seen'] or '—':<12}")

    if linkedin_only:
        print(f"\n--- ∅ LINKEDIN-ONLY (contacts found, 0 emails) ---")
        print(f"  {'Brand':<32} {'Identifier':<45} {'#C':>4}  {'Last seen':<12}")
        print("  " + "─" * 96)
        for name, identifier, info in linkedin_only:
            print(f"  {name:<32} {identifier:<45} {info['total']:>4}  {info['last_seen'] or '—':<12}")

    if empty:
        print(f"\n--- ✗ EMPTY (ZZ- files — Clay returned 0, excluded from next batch) ---")
        print(f"  {'Brand':<32} {'Identifier':<55}")
        print("  " + "─" * 90)
        for name, identifier in empty:
            print(f"  {name:<32} {identifier:<55}")

    if missing:
        print(f"\n--- ✗ MISSING (has identifier — never searched) ---")
        print(f"  {'Brand':<32} {'Identifier':<45}")
        print("  " + "─" * 80)
        for name, identifier in missing:
            print(f"  {name:<32} {identifier:<45}")

    if needs_domain:
        print(f"\n--- ? NEEDS-DOMAIN (no slug/website — web search required) ---")
        print(f"  {'Brand':<32}")
        print("  " + "─" * 34)
        for (name,) in needs_domain[:20]:
            print(f"  {name:<32}  [needs-domain]")
        if len(needs_domain) > 20:
            print(f"  ... and {len(needs_domain) - 20} more")

    # Next batch: missing first (have identifiers), then linkedin_only (retry for emails),
    # then needs_domain (require web search — listed last, agent handles via WebSearch).
    # Within each tier, Beauty industry brands come first.
    _is_beauty = lambda n: 0 if searchable.get(n, {}).get("industry", "").lower() == "beauty" else 1
    next_up  = sorted([(n, i, "missing")         for n, i    in missing],       key=lambda x: _is_beauty(x[0]))
    next_up += sorted([(n, i, "linkedin-only")   for n, i, _ in linkedin_only], key=lambda x: _is_beauty(x[0]))
    next_up += sorted([(n, None, "needs-domain") for (n,)    in needs_domain],  key=lambda x: _is_beauty(x[0]))

    if next_up:
        print(f"\n=== NEXT BATCH RECOMMENDATION (top 22) ===")
        print(f"  {'#':<4} {'Tier':<14} {'Brand':<32} {'Identifier'}")
        print("  " + "─" * 100)
        for idx, (name, identifier, tier) in enumerate(next_up[:22], 1):
            ident_str = identifier or "[needs-domain — use WebSearch]"
            print(f"  [{idx:>2}] {tier:<14}  {name:<32} {ident_str}")
    else:
        print("\nAll searchable brands are covered or pending LinkedIn-only retry.")



# ---------------------------------------------------------------------------
# BATCH — populated by Claude after each round of Clay MCP calls.
# Format: brand_name → (contacts_list, domain_hint)
# contacts_list items: {"name", "latest_experience_title", "url", "email", "domain"}
# ---------------------------------------------------------------------------
BATCH = {
    "Miu Miu": ([], "https://www.linkedin.com/company/miu-miu"),
    "Mixsoon": ([], "https://www.linkedin.com/company/mixsoon"),
    "Molton Brown": ([], "moltonbrown.co.uk"),
    "Mon Rêve Cosmetics": ([
        {"name": "Lydia Christoforidou", "latest_experience_title": "Brand Manager", "url": "https://www.linkedin.com/in/lydia-christoforidou-9278a4122/", "email": None, "domain": "hellenica.gr"},
        {"name": "Electra Papanastasiou", "latest_experience_title": "Marketing Manager, Strategy Lead Seventeen Cosmetics", "url": "https://www.linkedin.com/in/papanastasiou-electra/", "email": None, "domain": "hellenica.gr"},
        {"name": "Silia Karatza", "latest_experience_title": "Digital Marketing & Social Media Specialist", "url": "https://www.linkedin.com/in/silia-karatza-79b726197/", "email": None, "domain": "hellenica.gr"},
    ], "hellenica.gr"),
    "Moroccanoil": ([], "moroccanoil.com"),
    "Morphe": ([], "https://www.linkedin.com/company/morphe"),
    "Mugler": ([], "https://www.linkedin.com/company/mugler"),
    "My 7 Days": ([], "https://www.linkedin.com/company/7days-beauty"),
    "My7Days": ([], "https://www.linkedin.com/company/7days-beauty"),
    "NARS Cosmetics": ([], "narscosmetics.com"),
    "Natasha Denona": ([], "https://www.linkedin.com/company/natasha-denona"),
    "Natura Siberica": ([], "naturasiberica.ru"),
    "Neostrata": ([], "neostrata.com"),
    "Neutrogena": ([], "neutrogena.com"),
    "Notino": ([], "notino.com"),
    "Novexpert": ([], "novexpert-lab.fr"),
    "Nuxe": ([], "nuxe.com"),
    "OPI": ([], "https://www.linkedin.com/company/opi"),
    "Olaplex": ([], "https://www.linkedin.com/company/olaplex"),
    "Old Spice": ([], "https://www.linkedin.com/company/old-spice"),
    "Organic Shop": ([], "https://www.linkedin.com/company/organic-shop"),
    "Oriflame": ([
        {"name": "despoina noukari", "latest_experience_title": "oriflame marketing", "url": "https://www.linkedin.com/in/despoina-noukari-5a0b9aab/", "email": None, "domain": "oriflame.com"},
    ], "oriflame.com"),
}
_LAST_BATCH_2026_06_22 = {
    "Douleutaras": ([
        {"name": "Danai Dimitriou", "latest_experience_title": "Performance Marketing Specialist", "url": "https://www.linkedin.com/in/danai-dimitriou/", "email": None, "domain": "douleutaras.gr"},
        {"name": "Apostolos Chatziathanasiou", "latest_experience_title": "Performance Marketing Manager", "url": "https://www.linkedin.com/in/apostolos-chatziathanasiou-63a347180/", "email": None, "domain": "douleutaras.gr"},
        {"name": "Ermis Anastasopoulos", "latest_experience_title": "Communications Designer", "url": "https://www.linkedin.com/in/ermis-anastasopoulos/", "email": None, "domain": "douleutaras.gr"},
        {"name": "Maria Gkouti", "latest_experience_title": "Social Media Specialist", "url": "https://www.linkedin.com/in/maria-gkouti-8285b91ab/", "email": None, "domain": "douleutaras.gr"},
        {"name": "Maria Panagopoulou", "latest_experience_title": "Marketing Manager", "url": "https://www.linkedin.com/in/maria-panagopoulou-344315131/", "email": None, "domain": "douleutaras.gr"},
        {"name": "Christina Nicolaou", "latest_experience_title": "Social Media Specialist", "url": "https://www.linkedin.com/in/christina-nicolaou-6565981ba/", "email": None, "domain": "douleutaras.gr"},
    ], "douleutaras.gr"),

    "Box Now": ([
        {"name": "Chris Papandropoulos", "latest_experience_title": "Group CMO", "url": "https://www.linkedin.com/in/chrispapandropoulos/", "email": None, "domain": "boxnow.gr"},
        {"name": "Nikolaos Katsadramis", "latest_experience_title": "Marketing Executive", "url": "https://www.linkedin.com/in/nikolaos-katsadramis-185881129/", "email": None, "domain": "boxnow.gr"},
        {"name": "Anastasia Kalliaropoulou", "latest_experience_title": "Marketing Supervisor", "url": "https://www.linkedin.com/in/anastasia-kalliaropoulou-6572b2194/", "email": None, "domain": "boxnow.gr"},
        {"name": "Antonis Mpalamos", "latest_experience_title": "Director of Partners Operations", "url": "https://www.linkedin.com/in/antonis-mpalamos-2195825b/", "email": None, "domain": "boxnow.gr"},
        {"name": "ALEXANDROS VAGIAS", "latest_experience_title": "Sales Director", "url": "https://www.linkedin.com/in/alexandros-vagias-4811b27b/", "email": None, "domain": "boxnow.gr"},
    ], "boxnow.gr"),

    "Germanos": ([
        {"name": "Thanasis Panagopoulos", "latest_experience_title": "Digital and Social Media Senior Specialist", "url": "https://www.linkedin.com/in/thanasispanagopoulos/", "email": None, "domain": "germanos.gr"},
        {"name": "Maria Pitsiou", "latest_experience_title": "Corporate Communications COSMOTE eValue & Germanos", "url": "https://www.linkedin.com/in/maria-pitsiou-97310ba8/", "email": None, "domain": "germanos.gr"},
        {"name": "Georgia Konstantopoulou", "latest_experience_title": "Section Manager Trade Marketing GERMANOS Shops / BU RETAIL OTE GROUP", "url": "https://www.linkedin.com/in/georgia-konstantopoulou-b15768161/", "email": None, "domain": "germanos.gr"},
        {"name": "George Evangelinos", "latest_experience_title": "Brand Manager", "url": "https://www.linkedin.com/in/george-evangelinos-b771a9152/", "email": None, "domain": "germanos.gr"},
        {"name": "Christos Liaros", "latest_experience_title": "Trade Marketing", "url": "https://www.linkedin.com/in/christos-liaros-78752435/", "email": None, "domain": "germanos.gr"},
        {"name": "Chrisanthos Mavromatis", "latest_experience_title": "Sales Marketing Manager", "url": "https://www.linkedin.com/in/chrisanthos-mavromatis-721220162/", "email": None, "domain": "germanos.gr"},
        {"name": "lefteris bamos", "latest_experience_title": "Mobile Marketing Manager", "url": "https://www.linkedin.com/in/lefteris-bamos-0816922bb/", "email": None, "domain": "germanos.gr"},
        {"name": "George Grammatikopoulos", "latest_experience_title": "Marketing Team Member", "url": "https://www.linkedin.com/in/george-grammatikopoulos-467b42289/", "email": None, "domain": "germanos.gr"},
        {"name": "Spyros Kafalis", "latest_experience_title": "Trade Marketing Manager", "url": "https://www.linkedin.com/in/spyros-kafalis-6059bb13/", "email": None, "domain": "germanos.gr"},
    ], "germanos.gr"),

    "MEVGAL": ([
        {"name": "GIOTA BETZOUNI", "latest_experience_title": "Marketing Manager", "url": "https://www.linkedin.com/in/giota-betzouni-79355550/", "email": None, "domain": "mevgal.gr"},
        {"name": "Joanne Jouli", "latest_experience_title": "Marketing Manager", "url": "https://www.linkedin.com/in/joanne-jouli-17738418/", "email": None, "domain": "mevgal.gr"},
        {"name": "Stavroula Koutsou", "latest_experience_title": "Brand Manager", "url": "https://www.linkedin.com/in/stavroula-koutsou-05283211/", "email": None, "domain": "mevgal.gr"},
        {"name": "Stathis Manakos", "latest_experience_title": "Export Marketing Manager", "url": "https://www.linkedin.com/in/stathis-manakos-9a8a159a/", "email": None, "domain": "mevgal.gr"},
        {"name": "Stelios Malakopoulos", "latest_experience_title": "Brand Ambassador", "url": "https://www.linkedin.com/in/stelios-malakopoulos-3b8b1bb1/", "email": None, "domain": "mevgal.gr"},
        {"name": "George Vanidis", "latest_experience_title": "Exports Director", "url": "https://www.linkedin.com/in/george-vanidis-a974719/", "email": None, "domain": "mevgal.gr"},
        {"name": "Dimitrios Katsaras", "latest_experience_title": "Commercial director", "url": "https://www.linkedin.com/in/dimitrios-katsaras-95071698/", "email": None, "domain": "mevgal.gr"},
        {"name": "Konstantina Kokkinopoulou", "latest_experience_title": "Marketing Department", "url": "https://www.linkedin.com/in/konstantina-kokkinopoulou-aa693313b/", "email": None, "domain": "mevgal.gr"},
    ], "mevgal.gr"),

    "Vitex": ([
        {"name": "Eleni Souladaki", "latest_experience_title": "Trade Marketing Manager", "url": "https://www.linkedin.com/in/eleni-souladaki-15954b19/", "email": None, "domain": "vitex.gr"},
        {"name": "Vana Paraskevopoulou", "latest_experience_title": "Trade Marketing Specialist", "url": "https://www.linkedin.com/in/vana-paraskevopoulou-098a28204/", "email": None, "domain": "vitex.gr"},
        {"name": "Anastasis Nikolouzos", "latest_experience_title": "Trade Marketing Coordinator", "url": "https://www.linkedin.com/in/anastasis-nikolouzos-2943a52b3/", "email": None, "domain": "vitex.gr"},
        {"name": "Efthymios Koletsis", "latest_experience_title": "International Business Director", "url": "https://www.linkedin.com/in/efthymios-koletsis-527a805a/", "email": None, "domain": "vitex.gr"},
    ], "vitex.gr"),

    "Three Cents": ([
        {"name": "Alexander Sourmpatis", "latest_experience_title": "Global Brand Ambassador", "url": "https://www.linkedin.com/in/alexander-sourmpatis-4694811b/", "email": None, "domain": "threecents.com"},
        {"name": "George Bagos", "latest_experience_title": "Business Data Analyst", "url": "https://www.linkedin.com/in/george-bagos-2962216a/", "email": None, "domain": "threecents.com"},
        {"name": "Sophia Terzopoulou", "latest_experience_title": "Head of Marketing & Communications", "url": "https://www.linkedin.com/in/sophia-terzopoulou-a25a9340/", "email": None, "domain": "threecents.com"},
    ], "threecents.com"),

    "Snappi": ([
        {"name": "Dimitris Ganetsos", "latest_experience_title": "Head of Marketing | Snappi", "url": "https://www.linkedin.com/in/dimitris-ganetsos-43ab8226/", "email": None, "domain": "snappibank.com"},
        {"name": "Sofia Nefeli Spilioti", "latest_experience_title": "Senior Marketing Manager", "url": "https://www.linkedin.com/in/sofianefeli/", "email": None, "domain": "snappibank.com"},
        {"name": "Evita Vitsentzatou", "latest_experience_title": "Senior Marketing Manager", "url": "https://www.linkedin.com/in/evita-vitsentzatou-aaa95059/", "email": None, "domain": "snappibank.com"},
        {"name": "Piret Reinson", "latest_experience_title": "Head of Global Brand & Communications", "url": "https://www.linkedin.com/in/piret-reinson-03b12012/", "email": None, "domain": "snappibank.com"},
        {"name": "Marios Koumpas", "latest_experience_title": "Public Relations Manager", "url": "https://www.linkedin.com/in/marios-koumpas-07a8205a/", "email": None, "domain": "snappibank.com"},
        {"name": "Maria Mouzakiti", "latest_experience_title": "Head of Culture & Internal Communications", "url": "https://www.linkedin.com/in/mouzakiti/", "email": None, "domain": "snappibank.com"},
        {"name": "George Todoris", "latest_experience_title": "Digital Marketing Lead", "url": "https://www.linkedin.com/in/george-todoris-b2612029/", "email": None, "domain": "snappibank.com"},
        {"name": "Dimitra Tsigkri", "latest_experience_title": "Social Media & Content Coordinator", "url": "https://www.linkedin.com/in/dimitra-tsigkri/", "email": None, "domain": "snappibank.com"},
    ], "snappibank.com"),

    "LG": ([
        {"name": "John Mantas", "latest_experience_title": "Director, AirConditioning, Energy & B2B", "url": "https://www.linkedin.com/in/john-mantas-88b8091/", "email": None, "domain": "lg.com"},
        {"name": "Krisa Ylli", "latest_experience_title": "Marketing Supervisor (TV&Audio)", "url": "https://www.linkedin.com/in/krisa-ylli-b7b72a185/", "email": None, "domain": "lg.com"},
        {"name": "CHARALAMPOS NIKOLAOU", "latest_experience_title": "Digital Display & Project Implementation Specialist – B2B | LG Electronics (Greece)", "url": "https://www.linkedin.com/in/charalampos-nikolaou-9a248b191/", "email": None, "domain": "lg.com"},
    ], "lg.com"),
}

_LAST_BATCH = {
    "JYSK": ([
        {"name": "Kostas Dimopoulos DipWSET", "latest_experience_title": "Sales & Marketing Manager", "url": "https://www.linkedin.com/in/kostas-dimopoulos-dipwset-a2897320/", "email": None, "domain": "jysk.com"},
        {"name": "Eleni Kolotsiou", "latest_experience_title": "Commercial Sales and Marketing Assistant, Greece", "url": "https://www.linkedin.com/in/eleni-kolotsiou-6999682a/", "email": None, "domain": "jysk.com"},
        {"name": "Tonia Liadi", "latest_experience_title": "Jysk Influencer", "url": "https://www.linkedin.com/in/tonia-liadi-ba977b33b/", "email": None, "domain": "jysk.com"},
    ], "jysk.com"),

    "Kinder": ([
        {"name": "Elena Kosmatou", "latest_experience_title": "Brand Manager Kinder | GR/CY/MT", "url": "https://www.linkedin.com/in/elena-kosmatou-15556a11b/", "email": None, "domain": "ferrerocareers.com"},
        {"name": "Anna Kokkali", "latest_experience_title": "Brand Manager", "url": "https://www.linkedin.com/in/anna-kokkali-34497990/", "email": None, "domain": "ferrerocareers.com"},
        {"name": "Alexandros Tzagkarakis", "latest_experience_title": "Trade Marketing Manager", "url": "https://www.linkedin.com/in/alextzagkarakis/", "email": None, "domain": "ferrerocareers.com"},
        {"name": "Maria Pappa", "latest_experience_title": "Brand Manager | GCM", "url": "https://www.linkedin.com/in/maria-pappa-829085130/", "email": None, "domain": "ferrerocareers.com"},
        {"name": "Anthoula Flegka", "latest_experience_title": "Business Insights and Trade Marketing Specialist", "url": "https://www.linkedin.com/in/anthoula-flegka/", "email": None, "domain": "ferrerocareers.com"},
        {"name": "Dora Kontorouda", "latest_experience_title": "Country Marketing Manager - Nutella, Premium Chocolate, Biscuits, Candies", "url": "https://www.linkedin.com/in/dora-kontorouda-113b5b20/", "email": None, "domain": "ferrerocareers.com"},
        {"name": "Eftychia Dania", "latest_experience_title": "Brand Manager Kinder Greece, Cyprus, Malta", "url": "https://www.linkedin.com/in/eftychia-dania-9280448b/", "email": None, "domain": "ferrerocareers.com"},
        {"name": "Marianna Masoura", "latest_experience_title": "Country Marketing Manager", "url": "https://www.linkedin.com/in/marianna-masoura-900b5220b/", "email": None, "domain": "ferrerocareers.com"},
        {"name": "marina andronikou", "latest_experience_title": "marketing research manager", "url": "https://www.linkedin.com/in/marina-andronikou-42a4633a/", "email": None, "domain": "ferrerocareers.com"},
        {"name": "SAKIS BOURATZIS", "latest_experience_title": "Marketing", "url": "https://www.linkedin.com/in/sakis-bouratzis-380568100/", "email": None, "domain": "ferrerocareers.com"},
        {"name": "simone spada", "latest_experience_title": "MARKETING MANAGER", "url": "https://www.linkedin.com/in/simone-spada-3154778/", "email": None, "domain": "ferrerocareers.com"},
    ], "ferrerocareers.com"),

    "NIVEA": ([
        {"name": "Eleni Krietsepi", "latest_experience_title": "Senior Brand Manager", "url": "https://www.linkedin.com/in/elenikrietsepi/", "email": None, "domain": "beiersdorf.com"},
        {"name": "Arianna Gkiolia", "latest_experience_title": "Shopper and Customer Marketing Manager", "url": "https://www.linkedin.com/in/ariannagkiolia/", "email": None, "domain": "beiersdorf.com"},
        {"name": "Despoina Katsira", "latest_experience_title": "Brand Manager", "url": "https://www.linkedin.com/in/despoina-katsira-60622615b/", "email": None, "domain": "beiersdorf.com"},
        {"name": "Petros Chanis", "latest_experience_title": "Head of Shopper & Customer Marketing", "url": "https://www.linkedin.com/in/petros-chanis-a9796314/", "email": None, "domain": "beiersdorf.com"},
        {"name": "Marlene Tabet", "latest_experience_title": "Head of Marketing", "url": "https://www.linkedin.com/in/marlene-tabet-57853725/", "email": None, "domain": "beiersdorf.com"},
        {"name": "Lydia Karamaria", "latest_experience_title": "Junior Brand Manager", "url": "https://www.linkedin.com/in/lydiakaramaria/", "email": None, "domain": "beiersdorf.com"},
        {"name": "Marilena Ramoutsaki", "latest_experience_title": "Digital Marketing Specialist", "url": "https://www.linkedin.com/in/marilena-ramoutsaki/", "email": None, "domain": "beiersdorf.com"},
        {"name": "Alexandra Famelou", "latest_experience_title": "Media & PR Activation Manager", "url": "https://www.linkedin.com/in/alexandra-famelou-0a1734283/", "email": None, "domain": "beiersdorf.com"},
        {"name": "Efi Karanasiou", "latest_experience_title": "Medical Manager Europe & Marketing Manager Greece", "url": "https://www.linkedin.com/in/efi-karanasiou-b0a2665/", "email": None, "domain": "beiersdorf.com"},
        {"name": "Maria Ilektra Akarepi", "latest_experience_title": "Brand Manager", "url": "https://www.linkedin.com/in/maria-ilektra-akarepi-b8758112/", "email": None, "domain": "beiersdorf.com"},
        {"name": "Evi Christodoulou", "latest_experience_title": "Brand Manager", "url": "https://www.linkedin.com/in/evi-christodoulou-3b2044113/", "email": None, "domain": "beiersdorf.com"},
        {"name": "Elisavet Terzivassian", "latest_experience_title": "Jr Brand Manager HealthCare", "url": "https://www.linkedin.com/in/elisavet-terzivassian-a0b560257/", "email": None, "domain": "beiersdorf.com"},
    ], "beiersdorf.com"),

    "Dove": ([
        {"name": "Georgios Karavas", "latest_experience_title": "Senior Brand Manager Deodorants(AXE, Dove & Dove Men+Care)", "url": "https://www.linkedin.com/in/-george-karavas10/", "email": None, "domain": "unilever.com"},
        {"name": "Konstantina Dimoka", "latest_experience_title": "Senior Brand Manager Laundry Skip", "url": "https://www.linkedin.com/in/konstantina-dimoka-ab23b659/", "email": None, "domain": "unilever.com"},
        {"name": "Michaela Karpouzi", "latest_experience_title": "Head of Marketing GR & CY at UFS", "url": "https://www.linkedin.com/in/michaela-karpouzi/", "email": None, "domain": "unilever.com"},
        {"name": "Katerina Ntalli", "latest_experience_title": "Marketing Manager Skin Cleansing & Oral Care", "url": "https://www.linkedin.com/in/katerinantalli/", "email": None, "domain": "unilever.com"},
        {"name": "Kelly Karli", "latest_experience_title": "Marketing Manager Deodorants, Hair & Skin Care", "url": "https://www.linkedin.com/in/kelly-karli-9bb731bb/", "email": None, "domain": "unilever.com"},
        {"name": "Asylena Demerouti", "latest_experience_title": "Marketing Manager Skin & Oral Care", "url": "https://www.linkedin.com/in/asylena-demerouti-2904481a/", "email": None, "domain": "unilever.com"},
        {"name": "Valia Sakkou", "latest_experience_title": "Digital-Media-Commerce Strategy Lead (VET DMC) 1U Europe & Country Media Lead for Greece, Spain, CBB", "url": "https://www.linkedin.com/in/valia-sakkou-a1756174/", "email": None, "domain": "unilever.com"},
        {"name": "Aspa Diki", "latest_experience_title": "Marketing Manager Foods", "url": "https://www.linkedin.com/in/aspadiki/", "email": None, "domain": "unilever.com"},
        {"name": "Vicky Statha", "latest_experience_title": "Brand & Portfolio Manager", "url": "https://www.linkedin.com/in/vicky-statha-b371031b7/", "email": None, "domain": "unilever.com"},
        {"name": "Ευσταθία Παπαθανασίου", "latest_experience_title": "Senior Brand Manager Klinex Greece", "url": "https://www.linkedin.com/in/efstathia-papathanasiou/", "email": None, "domain": "unilever.com"},
        {"name": "KATERINA ALIFRAGKI", "latest_experience_title": "Trade Marketing Lead Homecare", "url": "https://www.linkedin.com/in/katerina-alifragki-08a34115/", "email": None, "domain": "unilever.com"},
        {"name": "Ioanna Lioliou", "latest_experience_title": "Head of Marketing", "url": "https://www.linkedin.com/in/ioannalioliou/", "email": None, "domain": "unilever.com"},
        {"name": "Evangelia Vossou", "latest_experience_title": "Marketing Team Lead Personal Care 1U countries Europe", "url": "https://www.linkedin.com/in/evangelia-vossou-b1271714/", "email": None, "domain": "unilever.com"},
        {"name": "Ilias Faratzis", "latest_experience_title": "Head of Shopper Marketing", "url": "https://www.linkedin.com/in/ilias-faratzis-22881774/", "email": None, "domain": "unilever.com"},
        {"name": "Silia Patsou", "latest_experience_title": "Home Care Business Unit Marketing Lead", "url": "https://www.linkedin.com/in/silia-patsou-89074b59/", "email": None, "domain": "unilever.com"},
        {"name": "Nikoletta Isari", "latest_experience_title": "Brand Manager Ice Cream", "url": "https://www.linkedin.com/in/nikoletta-isari/", "email": None, "domain": "unilever.com"},
        {"name": "Eleni Stella", "latest_experience_title": "Brand Manager Skin Cleansing", "url": "https://www.linkedin.com/in/eleni-stella/", "email": None, "domain": "unilever.com"},
        {"name": "Afrodite Foniadaki", "latest_experience_title": "Head of Marketing", "url": "https://www.linkedin.com/in/afrodite-foniadaki-577767a3/", "email": None, "domain": "unilever.com"},
        {"name": "Athanasios Vourdas", "latest_experience_title": "Business Development Manager", "url": "https://www.linkedin.com/in/athanasios-vourdas-75620685/", "email": None, "domain": "unilever.com"},
        {"name": "Eleni Papathanasiou", "latest_experience_title": "Assistant Brand Marketing Skip", "url": "https://www.linkedin.com/in/eleni-papathanasiou-b34baa196/", "email": None, "domain": "unilever.com"},
    ], "unilever.com"),

    "Converse": ([
        {"name": "Artemis Schistou", "latest_experience_title": "Senior Digital Merchandiser", "url": "https://www.linkedin.com/in/artemis-schistou-99827275/", "email": None, "domain": "converse.com"},
    ], "converse.com"),

    "BMW": ([
        {"name": "Dimitra Bikou", "latest_experience_title": "BMW Marketing Manager", "url": "https://www.linkedin.com/in/dimitra-bikou-7ab3a129/", "email": None, "domain": "bmwgroup.com"},
        {"name": "Alexandra Karvouni", "latest_experience_title": "Marketing & Project Specialist", "url": "https://www.linkedin.com/in/alexandra-karvouni-192030130/", "email": None, "domain": "bmwgroup.com"},
        {"name": "Michael Gaganis", "latest_experience_title": "MINI Marketing Manager", "url": "https://www.linkedin.com/in/michael-gaganis-9a439b7/", "email": None, "domain": "bmwgroup.com"},
        {"name": "Irene-Penelope Angeletou", "latest_experience_title": "Digital & CRM Specialist", "url": "https://www.linkedin.com/in/ireneangeletou/", "email": None, "domain": "bmwgroup.com"},
        {"name": "Katerina Ninou", "latest_experience_title": "Customer Support Marketing & Projects Specialist", "url": "https://www.linkedin.com/in/κατερίνα-νίνου/", "email": None, "domain": "bmwgroup.com"},
        {"name": "Elli Stasinou", "latest_experience_title": "Events Expert & Dealer Marketing Coordinator", "url": "https://www.linkedin.com/in/elli-stasinou-b7b43559/", "email": None, "domain": "bmwgroup.com"},
        {"name": "Dimitris Kokalias", "latest_experience_title": "BMW Motorrad Sales and Marketing Specialist", "url": "https://www.linkedin.com/in/dimitris-kokalias-860a38271/", "email": None, "domain": "bmwgroup.com"},
        {"name": "Elpida Sella", "latest_experience_title": "Motorrad Marketing and Sales Specialist", "url": "https://www.linkedin.com/in/elpida-sella-73bb45279/", "email": None, "domain": "bmwgroup.com"},
        {"name": "Lisa Ceruvija", "latest_experience_title": "Senior Marketing Specialist", "url": "https://www.linkedin.com/in/lisa-ceruvija-859b09246/", "email": None, "domain": "bmwgroup.com"},
    ], "bmwgroup.com"),

    "Ferryhopper": ([
        {"name": "Kaj van Zweeden", "latest_experience_title": "Chief Marketing Officer", "url": "https://www.linkedin.com/in/kaj-van-zweeden-43a552b6/", "email": None, "domain": "ferryhopper.com"},
        {"name": "Sofia Drogoudi", "latest_experience_title": "Design Lead [Marketing]", "url": "https://www.linkedin.com/in/sofiadrogoudi/", "email": None, "domain": "ferryhopper.com"},
        {"name": "Athina Sofou", "latest_experience_title": "Performance Marketing Manager", "url": "https://www.linkedin.com/in/athinasofou/", "email": None, "domain": "ferryhopper.com"},
        {"name": "Christina Gkini", "latest_experience_title": "Head of Performance Marketing", "url": "https://www.linkedin.com/in/christina-gkini/", "email": None, "domain": "ferryhopper.com"},
        {"name": "Michalis Kakotaritis", "latest_experience_title": "Brand Lead", "url": "https://www.linkedin.com/in/michalis-kakotaritis-88844493/", "email": None, "domain": "ferryhopper.com"},
        {"name": "Aristeidis Remoundos", "latest_experience_title": "Affiliate Marketing Specialist", "url": "https://www.linkedin.com/in/aristeidis-remoundos-49a2701ba/", "email": None, "domain": "ferryhopper.com"},
        {"name": "Nicole Adonopoulos", "latest_experience_title": "Brand Specialist", "url": "https://www.linkedin.com/in/nicole-adonopoulos-5683b462/", "email": None, "domain": "ferryhopper.com"},
        {"name": "Andromachi Ferro", "latest_experience_title": "Marketing Expert", "url": "https://www.linkedin.com/in/andromachi-ferro-55409016b/", "email": None, "domain": "ferryhopper.com"},
        {"name": "Chris Michalopoulos", "latest_experience_title": "Creative Marketing Specialist", "url": "https://www.linkedin.com/in/chris-michalopoulos/", "email": None, "domain": "ferryhopper.com"},
        {"name": "Ioannis Floros", "latest_experience_title": "Social Media Marketing Specialist", "url": "https://www.linkedin.com/in/ioannis-floros-b9b5681a9/", "email": None, "domain": "ferryhopper.com"},
    ], "ferryhopper.com"),

    "Vans": ([
        {"name": "Labrini Kranioti", "latest_experience_title": "Marketing Executive", "url": "https://www.linkedin.com/in/labrinikranioti/", "email": None, "domain": "vans.com"},
    ], "vans.com"),
}

_LAST_BATCH = {
    "ANT1": ([
        {"name": "Konstantinos Bourounis", "latest_experience_title": "Chief Marketing Officer - Antenna Group Greece", "url": "https://www.linkedin.com/in/konstantinos-bourounis-8675a0/", "email": None, "domain": "ant1.gr"},
        {"name": "Tasos Katsikadakos", "latest_experience_title": "Commercial, Digital & Research Director", "url": "https://www.linkedin.com/in/tasos-katsikadakos-4566856/", "email": None, "domain": "ant1.gr"},
        {"name": "Eva Vazaka", "latest_experience_title": "Commercial Manager | ANT1 Events & Sponsorships", "url": "https://www.linkedin.com/in/evazaka/", "email": None, "domain": "ant1.gr"},
        {"name": "Olympia Tsamasfyra", "latest_experience_title": "Marketing Manager", "url": "https://www.linkedin.com/in/olympia-tsamasfyra-40033069/", "email": None, "domain": "ant1.gr"},
        {"name": "Nancy Fafouti", "latest_experience_title": "Social Media Manager", "url": "https://www.linkedin.com/in/nancy-fafouti/", "email": None, "domain": "ant1.gr"},
        {"name": "Agapi Kantartzi", "latest_experience_title": "Marketing Manager & Communications | Antenna Audio / easy 97.2 | Ρυθμος 94.9 | Soundis.gr", "url": "https://www.linkedin.com/in/agapi/", "email": None, "domain": "ant1.gr"},
        {"name": "Thetis Gouliou", "latest_experience_title": "Chief Strategy & Business Development Officer", "url": "https://www.linkedin.com/in/thetis-gouliou-63680934/", "email": None, "domain": "ant1.gr"},
        {"name": "Tina Magaraki", "latest_experience_title": "Head of Business Development", "url": "https://www.linkedin.com/in/tina-magaraki-46b8086/", "email": None, "domain": "ant1.gr"},
        {"name": "Savvas Vitalis", "latest_experience_title": "Head of Digital Sales", "url": "https://www.linkedin.com/in/savvas-vitalis-00082b27/", "email": None, "domain": "ant1.gr"},
        {"name": "Kiki Malerdou", "latest_experience_title": "Commercial Manager Radio & Events", "url": "https://www.linkedin.com/in/kiki-malerdou-8803742a/", "email": None, "domain": "ant1.gr"},
        {"name": "Iraklis Ioannidis", "latest_experience_title": "Music Marketing Lead - Marketing Manager Heaven/Warner & Antenna Intelligence", "url": "https://www.linkedin.com/in/iraklis-ioannidis-0a536622/", "email": None, "domain": "ant1.gr"},
        {"name": "Ioanna Panagioti", "latest_experience_title": "Social Media Expert", "url": "https://www.linkedin.com/in/ioanna-panagioti-25bba91a9/", "email": None, "domain": "ant1.gr"},
        {"name": "Vlasia Manteska", "latest_experience_title": "Digital Sales Executive", "url": "https://www.linkedin.com/in/vlasia-manteska-848032140/", "email": None, "domain": "ant1.gr"},
        {"name": "Marco Struecker", "latest_experience_title": "Interim General Manager - ANT1+", "url": "https://www.linkedin.com/in/marcostruecker/", "email": None, "domain": "ant1.gr"},
    ], "ant1.gr"),

    "AEK FC": ([
        {"name": "Antonis Apostolopoulos", "latest_experience_title": "Event Management & Sports Marketing Assistant", "url": "https://www.linkedin.com/in/antonis-apostolopoulos-7b753b187/", "email": None, "domain": "aekfc.gr"},
        {"name": "Christina Koromila", "latest_experience_title": "Social Media", "url": "https://www.linkedin.com/in/christina-koromila-768326b5/", "email": None, "domain": "aekfc.gr"},
    ], "aekfc.gr"),

    "Jumbo": ([
        {"name": "Andreas Xenofontos", "latest_experience_title": "Head of Digital Marketing & E-commerce", "url": "https://www.linkedin.com/in/andreas-xenofontos/", "email": None, "domain": "e-jumbo.gr"},
        {"name": "Maria Gkara", "latest_experience_title": "Media and Communications Consultant", "url": "https://www.linkedin.com/in/maria-gkara-7940535/", "email": None, "domain": "e-jumbo.gr"},
        {"name": "Andra Bilici", "latest_experience_title": "Digital Marketing & E-commerce Manager", "url": "https://www.linkedin.com/in/andra-bilici-731278144/", "email": None, "domain": "e-jumbo.gr"},
        {"name": "Panagiotis Mylonidis", "latest_experience_title": "E-Commerce & Social Media Manager", "url": "https://www.linkedin.com/in/pmylonidis/", "email": None, "domain": "e-jumbo.gr"},
    ], "e-jumbo.gr"),

    "FAGE": ([
        {"name": "Katerina Gkouvaki", "latest_experience_title": "Senior Brand Manager GR", "url": "https://www.linkedin.com/in/katerina-gkouvaki-743462183/", "email": None, "domain": "home.fage"},
        {"name": "Alexis Alexopoulos", "latest_experience_title": "Marketing & Communications Director", "url": "https://www.linkedin.com/in/alexis-alexopoulos-8a539b1/", "email": None, "domain": "home.fage"},
    ], "home.fage"),

    "New Balance": ([
        {"name": "Eleana Ravani", "latest_experience_title": "Trade & Marketing Communication", "url": "https://www.linkedin.com/in/eleana-ravani-917520193/", "email": None, "domain": "newbalance.com"},
        {"name": "Panagiotis Kyvrikosaios", "latest_experience_title": "Retail Marketing and Running Category Manager", "url": "https://www.linkedin.com/in/panoskyvrikosaios/", "email": None, "domain": "newbalance.com"},
    ], "newbalance.com"),

    "Zara": ([
        {"name": "Mary Geroukali", "latest_experience_title": "Communications - PR & Marketing Director (Greece)", "url": "https://www.linkedin.com/in/mary-geroukali-05250731/", "email": None, "domain": "inditexpeople.com"},
    ], "inditexpeople.com"),

    "Puma": ([
        {"name": "Esteve Planas", "latest_experience_title": "General Manager PUMA Southern Europe", "url": "https://www.linkedin.com/in/esteveplanas/", "email": None, "domain": "puma.com"},
        {"name": "Pedro Moscoso", "latest_experience_title": "Marketing & Communications Director Southern Europe", "url": "https://www.linkedin.com/in/pedromoscoso/", "email": None, "domain": "puma.com"},
        {"name": "Carmen Ponce", "latest_experience_title": "Lead Brand Consumer Southern Europe", "url": "https://www.linkedin.com/in/carmenponce6/", "email": None, "domain": "puma.com"},
    ], "puma.com"),

    "CarVertical": ([
        {"name": "Mantas Ribelis", "latest_experience_title": "Chief Marketing Officer", "url": "https://www.linkedin.com/in/mantasribelis/", "email": None, "domain": "carvertical.com"},
        {"name": "Ovidijus S.", "latest_experience_title": "Influencer Marketing Performance Manager", "url": "https://www.linkedin.com/in/ovidijussaldauskas/", "email": None, "domain": "carvertical.com"},
        {"name": "Tautvydas Sirevicius", "latest_experience_title": "Influencer Marketing Manager", "url": "https://www.linkedin.com/in/tautvydas-sireviäius/", "email": None, "domain": "carvertical.com"},
        {"name": "Deivydas Simas", "latest_experience_title": "Influencer Marketing Strategist", "url": "https://www.linkedin.com/in/deivydas-å¡imas/", "email": None, "domain": "carvertical.com"},
        {"name": "Gerda Servaite", "latest_experience_title": "Influencer Marketing Manager", "url": "https://www.linkedin.com/in/gerda-servait%c4%97-647774151/", "email": None, "domain": "carvertical.com"},
        {"name": "Ieva Pauliukoniene", "latest_experience_title": "Manager Strategic Partnerships", "url": "https://www.linkedin.com/in/ievapaul/", "email": None, "domain": "carvertical.com"},
        {"name": "Daniel Artisiuk", "latest_experience_title": "Senior Country Manager | B2B Sales & Partnerships, EU Markets", "url": "https://www.linkedin.com/in/danielartisiuk/", "email": None, "domain": "carvertical.com"},
    ], "carvertical.com"),
}


def git_sync():
    """Stage employee files + actionable_contacts.json, commit, pull --rebase, push."""
    files = [
        "AI Sales Agent System/actionable_contacts.json",
        "AI Sales Agent System/output/employees/",
    ]
    subprocess.run(["git", "-C", str(ROOT), "add"] + files, check=True)
    # Nothing staged → nothing to do
    diff = subprocess.run(["git", "-C", str(ROOT), "diff", "--cached", "--quiet"])
    if diff.returncode == 0:
        print("git_sync: nothing to commit — store unchanged.")
        return
    msg = f"clay-discoverer: batch save {TODAY}"
    subprocess.run(["git", "-C", str(ROOT), "commit", "-m", msg], check=True)
    subprocess.run(["git", "-C", str(ROOT), "pull", "--rebase"], check=True)
    subprocess.run(["git", "-C", str(ROOT), "push"], check=True)
    print(f"git_sync: pushed — '{msg}'")


def export_excel():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not installed. Run: pip3 install openpyxl --break-system-packages")
        return

    GREEN  = PatternFill("solid", fgColor="C6EFCE")  # covered (has email)
    YELLOW = PatternFill("solid", fgColor="FFEB9C")  # linkedin-only (no email)
    RED    = PatternFill("solid", fgColor="FFC7CE")  # empty (ZZ-)
    HEADER = PatternFill("solid", fgColor="2F4F7F")  # header row

    bold_white = Font(bold=True, color="FFFFFF")
    bold_black = Font(bold=True)
    center     = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left       = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin       = Side(style="thin", color="CCCCCC")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()
    ws = wb.active
    ws.title = "Contacts"

    headers = ["Brand", "Status", "Name", "Job Title", "Email", "LinkedIn URL", "Domain", "First Seen", "Last Seen"]
    col_widths = [28, 14, 28, 36, 34, 52, 22, 12, 12]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill   = HEADER
        cell.font   = bold_white
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    store = _scan_store()
    data  = json.loads(CONTACTS_FILE.read_text(encoding="utf-8"))
    all_brands = sorted(b["brand"] for b in data["brands"])

    row = 2
    for brand in all_brands:
        info = store.get(brand)
        if info is None:
            continue  # never searched — skip from export

        prefix = info["_prefix"]
        if prefix == "ZZ-":
            fill   = RED
            status = "Empty"
            contacts_list = []
        elif prefix == "00-":
            fill   = YELLOW
            status = "LinkedIn-only"
            f = STORE_DIR / f"00-{brand}.json"
            contacts_list = json.loads(f.read_text(encoding="utf-8")) if f.exists() else []
        else:
            fill   = GREEN
            status = "Covered"
            f = STORE_DIR / f"{brand}.json"
            contacts_list = json.loads(f.read_text(encoding="utf-8")) if f.exists() else []

        if not contacts_list:
            # One placeholder row for brands with no contacts
            vals = [brand, status, "—", "—", "—", "—", "—", "—", "—"]
            for col, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col, value=v)
                cell.fill = fill
                cell.alignment = left
                cell.border = border
                if col <= 2:
                    cell.font = bold_black
            row += 1
        else:
            for c in contacts_list:
                vals = [
                    brand, status,
                    c.get("name") or "—",
                    c.get("job_title") or "—",
                    c.get("email") or "—",
                    c.get("linkedin_url") or "—",
                    c.get("domain") or "—",
                    c.get("first_seen") or "—",
                    c.get("last_seen") or "—",
                ]
                for col, v in enumerate(vals, 1):
                    cell = ws.cell(row=row, column=col, value=v)
                    cell.fill = fill
                    cell.alignment = left
                    cell.border = border
                    if col <= 2:
                        cell.font = bold_black
                row += 1

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    wb.save(EXPORT_PATH)
    print(f"Exported {row - 2} rows to: {EXPORT_PATH}")
    print(f"  Green  = Covered (has email)")
    print(f"  Yellow = LinkedIn-only (no email yet)")
    print(f"  Red    = Empty (Clay returned 0 contacts)")


def list_for_enrichment(brand_names):
    """Print JSON array of {contactName, companyIdentifier} for contacts missing emails.

    Paste the output directly into find-and-enrich-list-of-contacts with
    dataPoints: {contactDataPoints: [{type: "Email"}]}.
    """
    result = []
    for brand in brand_names:
        path = _employee_path(brand)
        if not path.exists():
            print(f"# {brand}: no employee file — run Phase 1 first", file=sys.stderr)
            continue
        contacts = json.loads(path.read_text(encoding="utf-8"))
        pending = [c for c in contacts if not c.get("email")]
        if not pending:
            print(f"# {brand}: all contacts already have emails", file=sys.stderr)
            continue
        before = len(result)
        for c in pending:
            domain = c.get("domain") or DOMAIN_MAP.get(brand)
            if not domain:
                print(f"# {brand} / {c['name']}: no domain — add to DOMAIN_MAP", file=sys.stderr)
                continue
            result.append({"contactName": c["name"], "companyIdentifier": domain})
        print(f"# {brand}: {len(result) - before}/{len(pending)} contacts queued for email enrichment", file=sys.stderr)
    print(json.dumps(result, indent=2, ensure_ascii=False))


def purge_banned():
    """Remove contacts with banned titles from every file in the store."""
    total_removed = 0
    for f in sorted(STORE_DIR.glob("*.json")):
        if f.name == "desktop.ini" or f.stem.startswith("ZZ-"):
            continue
        contacts = json.loads(f.read_text(encoding="utf-8"))
        clean = [c for c in contacts if not _is_banned(c.get("job_title") or "")]
        removed = len(contacts) - len(clean)
        if removed:
            f.write_text(json.dumps(clean, indent=2, ensure_ascii=False), encoding="utf-8")
            print(f"  {f.name}: removed {removed} contact(s) with banned title(s)")
            total_removed += removed
    print(f"Purge complete — {total_removed} contact(s) removed across the store.")


if __name__ == "__main__":
    if "git_sync" in sys.argv:
        git_sync()
    elif "save" in sys.argv:
        if not BATCH:
            print("BATCH is empty — nothing to save.")
        else:
            for brand, (contacts, domain) in BATCH.items():
                save_contacts(brand, contacts, domain)
            print("Done.")
    elif "enrich_emails" in sys.argv:
        idx = sys.argv.index("enrich_emails")
        brands = sys.argv[idx + 1:]
        if not brands:
            print("Usage: python discover_contacts.py enrich_emails \"Brand1\" [\"Brand2\" ...]")
        else:
            list_for_enrichment(brands)
    elif "purge" in sys.argv:
        purge_banned()
    elif "export" in sys.argv:
        export_excel()
    elif "mark_empty" in sys.argv:
        idx = sys.argv.index("mark_empty")
        brands = sys.argv[idx + 1:]
        if not brands:
            print("Usage: python discover_contacts.py mark_empty \"Brand1\" \"Brand2\" ...")
        else:
            for b in brands:
                zz = STORE_DIR / f"ZZ-{b}.json"
                zz.write_text("[]", encoding="utf-8")
                old = STORE_DIR / f"00-{b}.json"
                if old.exists():
                    old.unlink()
                _sync_actionable(b)
                print(f"  '{b}' → ZZ-{b}.json (marked empty)")
    elif "unmark_empty" in sys.argv:
        idx = sys.argv.index("unmark_empty")
        brands = sys.argv[idx + 1:]
        if not brands:
            print("Usage: python discover_contacts.py unmark_empty \"Brand1\" ...")
        else:
            for b in brands:
                zz = STORE_DIR / f"ZZ-{b}.json"
                if zz.exists():
                    zz.unlink()
                    _sync_actionable(b)
                    print(f"  '{b}' removed from empty — will appear in next batch")
                else:
                    print(f"  '{b}' was not marked empty")
    else:
        status_report()
